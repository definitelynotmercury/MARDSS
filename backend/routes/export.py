from flask import Blueprint, jsonify, request, send_file
import mysql.connector
from config import DB_CONFIG, client
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
import json
from flask import request as flask_request
from auth_utils import token_required

export_bp = Blueprint('export', __name__)
def get_db():
    return mysql.connector.connect(**DB_CONFIG)


@export_bp.route('/api/export/data')
@token_required
def get_export_data():
    year_from = request.args.get('year_from', 'ALL')
    year_to = request.args.get('year_to', 'ALL')
    municipality = request.args.get('municipality', 'ALL')
    type_ = request.args.get('type', 'ALL')
    sections = request.args.get('sections', '').split(',')

    filters = []
    params = []

    if year_from != 'ALL' and year_to != 'ALL':
        filters.append("r.year BETWEEN %s AND %s")
        params.append(year_from)
        params.append(year_to)
    if municipality != 'ALL':
        filters.append("m.municipality_name = %s")
        params.append(municipality)
    if type_ != 'ALL':
        filters.append("a.type_name = %s")
        params.append(type_)

    where_clause = "WHERE " + " AND ".join(filters) if filters else ""

    result = {}

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    if 'dashboardKpi' in sections:
        cursor.execute(f"""
            SELECT SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where_clause}
        """, params)
        sum_req = cursor.fetchone()

        cursor.execute(f"""
            SELECT a.type_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where_clause}
            GROUP BY r.assistance_type_id
            ORDER BY total DESC
            LIMIT 1
        """, params)
        top_type = cursor.fetchone()

        cursor.execute(f"""
            SELECT m.municipality_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where_clause}
            GROUP BY r.municipality_id
            ORDER BY total DESC
            LIMIT 1
        """, params)
        top_municipality = cursor.fetchone()

        result['kpi'] = {
            "total_requests": sum_req["total"] or 0,
            "top_type": top_type or {"type_name": "N/A", "total": 0},
            "top_municipality": top_municipality or {"municipality_name": "N/A", "total": 0}
        }
    if 'dashboardTrends' in sections:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT r.year, a.type_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            {where_clause}
            GROUP BY r.year, r.assistance_type_id
            ORDER BY r.year
        """, params)
        trends = cursor.fetchall()

        result['trends'] = trends
        
    cursor.close()
    conn.close()        
    return jsonify(result)

@export_bp.route('/api/export/dataset')
@token_required
def get_dataset():
    year_from = request.args.get('year_from', 'ALL')
    year_to = request.args.get('year_to', 'ALL')
    municipality = request.args.get('municipality', 'ALL')
    type_ = request.args.get('type', 'ALL')
    month = request.args.get('month', 'ALL')

    filters = []
    params = []

    if year_from != 'ALL' and year_to != 'ALL':
        filters.append("r.year BETWEEN %s AND %s")
        params.append(year_from)
        params.append(year_to)
    if municipality != 'ALL':
        filters.append("m.municipality_name = %s")
        params.append(municipality)
    if type_ != 'ALL':
        filters.append("a.type_name = %s")
        params.append(type_)
    if month != 'ALL':
        filters.append("r.month = %s")
        params.append(month)

    where_clause = "WHERE " + " AND ".join(filters) if filters else ""

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(f"""
        SELECT r.year, r.month, m.municipality_name, a.type_name, r.request_count
        FROM assistance_records r
        JOIN assistance_types a ON r.assistance_type_id = a.type_id
        JOIN municipalities m ON r.municipality_id = m.municipality_id
        {where_clause}
        ORDER BY r.year, r.month, m.municipality_name, a.type_name
    """, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    return jsonify(data)


@export_bp.route('/api/export/dataset/excel')
@token_required
def export_dataset_excel():
    year_from = request.args.get('year_from', 'ALL')
    year_to = request.args.get('year_to', 'ALL')
    municipality = request.args.get('municipality', 'ALL')
    type_ = request.args.get('type', 'ALL')
    month = request.args.get('month', 'ALL')

    filters = []
    params = []

    if year_from != 'ALL' and year_to != 'ALL':
        filters.append("r.year BETWEEN %s AND %s")
        params.extend([year_from, year_to])
    if municipality != 'ALL':
        filters.append("m.municipality_name = %s")
        params.append(municipality)
    if type_ != 'ALL':
        filters.append("a.type_name = %s")
        params.append(type_)
    if month != 'ALL':
        filters.append("r.month = %s")
        params.append(month)

    where_clause = "WHERE " + " AND ".join(filters) if filters else ""

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    # Fixed row/column order — must match the DB seed order, which matches the upload template
    cursor.execute("SELECT municipality_name FROM municipalities ORDER BY municipality_id")
    all_municipalities = [row['municipality_name'] for row in cursor.fetchall()]

    cursor.execute("SELECT type_name FROM assistance_types ORDER BY type_id")
    all_types = [row['type_name'] for row in cursor.fetchall()]

    cursor.execute(f"""
        SELECT r.year, r.month, m.municipality_name, a.type_name, r.request_count
        FROM assistance_records r
        JOIN assistance_types a ON r.assistance_type_id = a.type_id
        JOIN municipalities m ON r.municipality_id = m.municipality_id
        {where_clause}
        ORDER BY r.year, r.month
    """, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    month_names = {
        1: 'JANUARY', 2: 'FEBRUARY', 3: 'MARCH', 4: 'APRIL', 5: 'MAY', 6: 'JUNE',
        7: 'JULY', 8: 'AUGUST', 9: 'SEPTEMBER', 10: 'OCTOBER', 11: 'NOVEMBER', 12: 'DECEMBER'
    }

    # Group records by (year, month) so each period becomes its own sheet
    by_period = {}
    for row in data:
        key = (row['year'], row['month'])
        by_period.setdefault(key, {})[(row['municipality_name'], row['type_name'])] = row['request_count']

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")

    for (year, month_num), lookup in sorted(by_period.items()):
        sheet_name = f"{month_names.get(month_num, month_num)} {year}"[:31]  # Excel caps sheet names at 31 chars
        ws = wb.create_sheet(sheet_name)

        # Header row: blank corner cell, then assistance types, then TOTAL
        ws.cell(row=1, column=1, value=' ')
        for col_idx, type_name in enumerate(all_types, start=2):
            ws.cell(row=1, column=col_idx, value=type_name)
        total_col = len(all_types) + 2
        ws.cell(row=1, column=total_col, value='TOTAL')
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Data rows: one per municipality
        last_data_row = len(all_municipalities) + 1
        for row_idx, muni in enumerate(all_municipalities, start=2):
            ws.cell(row=row_idx, column=1, value=muni)
            for col_idx, type_name in enumerate(all_types, start=2):
                count = lookup.get((muni, type_name), 0)
                ws.cell(row=row_idx, column=col_idx, value=count)
            first_letter = openpyxl.utils.get_column_letter(2)
            last_type_letter = openpyxl.utils.get_column_letter(total_col - 1)
            ws.cell(row=row_idx, column=total_col,
                    value=f"=SUM({first_letter}{row_idx}:{last_type_letter}{row_idx})")

        # TOTAL row at the bottom
        total_row = last_data_row + 1
        ws.cell(row=total_row, column=1, value='TOTAL')
        for col_idx in range(2, total_col + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            ws.cell(row=total_row, column=col_idx,
                    value=f"=SUM({col_letter}2:{col_letter}{last_data_row})")
        for cell in ws[total_row]:
            cell.font = Font(bold=True)

        ws.column_dimensions['A'].width = 22
        for col_idx in range(2, total_col + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 14

    if not wb.sheetnames:
        ws = wb.create_sheet("No Data")
        ws['A1'] = "No data found for selected filters."

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='MARDSS_Dataset.xlsx'
    )

@export_bp.route('/api/export/dataset/pdf')
@token_required
def export_dataset_pdf():
    year_from = request.args.get('year_from', 'ALL')
    year_to = request.args.get('year_to', 'ALL')
    municipality = request.args.get('municipality', 'ALL')
    type_ = request.args.get('type', 'ALL')

    filters = []
    params = []

    if year_from != 'ALL' and year_to != 'ALL':
        filters.append("r.year BETWEEN %s AND %s")
        params.extend([year_from, year_to])
    if municipality != 'ALL':
        filters.append("m.municipality_name = %s")
        params.append(municipality)
    if type_ != 'ALL':
        filters.append("a.type_name = %s")
        params.append(type_)

    where_clause = "WHERE " + " AND ".join(filters) if filters else ""

    conn = get_db()
    cursor = conn.cursor(dictionary=True)
    cursor.execute(f"""
        SELECT r.year, m.municipality_name, a.type_name, r.request_count
        FROM assistance_records r
        JOIN assistance_types a ON r.assistance_type_id = a.type_id
        JOIN municipalities m ON r.municipality_id = m.municipality_id
        {where_clause}
        ORDER BY r.year, m.municipality_name, a.type_name
    """, params)
    data = cursor.fetchall()
    cursor.close()
    conn.close()

    muni_label = 'All Municipalities' if municipality == 'ALL' else municipality
    type_label = 'All Types' if type_ == 'ALL' else type_

    # Setup
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    blue = colors.HexColor('#1D4ED8')
    light_blue = colors.HexColor('#EFF6FF')
    white = colors.white

    title_style = ParagraphStyle(
        'Title',
        parent=styles['Normal'],
        fontSize=15,
        fontName='Helvetica-Bold',
        textColor=white,
        alignment=TA_CENTER,
        spaceAfter=2
    )
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=9,
        fontName='Helvetica',
        textColor=white,
        alignment=TA_CENTER,
        spaceAfter=2
    )

    elements = []

    # Header banner
    header_data = [[
        Paragraph("MARDSS REPORT", title_style),
    ]]
    header_table = Table(header_data, colWidths=[26*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), blue),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('ROUNDEDCORNERS', [6, 6, 6, 6]),
    ]))
    elements.append(header_table)

    # Subtitle banner
    sub_data = [[
        Paragraph("Medical Assistance Request Decision Support System", subtitle_style),
    ],[
        Paragraph(f"Province of Bulacan - PSWDO | Period: {year_from} – {year_to}", subtitle_style),
    ],[
        Paragraph(f"Municipality: {muni_label} | Assistance Type: {type_label}", subtitle_style),
    ]]
    sub_table = Table(sub_data, colWidths=[26*cm])
    sub_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), blue),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('TOPPADDING', (0,0), (-1,-1), 0),
    ]))
    elements.append(sub_table)
    elements.append(Spacer(1, 0.5*cm))

    # Table header + data
    col_headers = ['Year', 'Municipality', 'Assistance Type', 'Request Count']
    table_data = [col_headers] + [
        [row['year'], row['municipality_name'], row['type_name'], row['request_count']]
        for row in data
    ]

    col_widths = [4*cm, 8*cm, 10*cm, 4*cm]
    main_table = Table(table_data, colWidths=col_widths, repeatRows=1)

    row_count = len(table_data)
    style_commands = [
        # Header row
        ('BACKGROUND', (0,0), (-1,0), blue),
        ('TEXTCOLOR', (0,0), (-1,0), white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,0), 10),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('BOTTOMPADDING', (0,0), (-1,0), 8),
        ('TOPPADDING', (0,0), (-1,0), 8),
        # Data rows
        ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
        ('FONTSIZE', (0,1), (-1,-1), 9),
        ('ALIGN', (3,1), (3,-1), 'RIGHT'),
        ('TOPPADDING', (0,1), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        # Borders
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#BFDBFE')),
        ('LINEBELOW', (0,0), (-1,0), 1.5, blue),
    ]

    # Alternating row colors
    for i in range(1, row_count):
        if i % 2 == 0:
            style_commands.append(('BACKGROUND', (0,i), (-1,i), light_blue))
        else:
            style_commands.append(('BACKGROUND', (0,i), (-1,i), white))

    main_table.setStyle(TableStyle(style_commands))
    elements.append(main_table)

    doc.build(elements)
    output.seek(0)

    filename = f"MARDSS_{year_from}-{year_to}_{muni_label}_{type_label}.pdf"
    return send_file(
        output,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@export_bp.route('/api/export/charts/excel', methods=['POST'])
@token_required
def export_charts_excel():
    body = flask_request.get_json()
    selected_sections = body.get('sections', [])
    filters = body.get('filters', {})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    blue = "1D4ED8"
    light_blue = "EFF6FF"

    def style_header_row(ws, row_num):
        for cell in ws[row_num]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=blue)
            cell.alignment = Alignment(horizontal='center')

    def style_data_rows(ws, start_row):
        for i, row in enumerate(ws.iter_rows(min_row=start_row)):
            for cell in row:
                if i % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=light_blue)

    conn = get_db()
    cursor = conn.cursor(dictionary=True)

    # ── Dashboard KPI ──
    if 'dashboardKpi' in selected_sections:
        f = filters.get('dashboardKpi', {})
        year = f.get('year', 'ALL')
        municipality = f.get('municipality', 'ALL')
        type_ = f.get('type', 'ALL')

        conditions, params = [], []
        if year != 'ALL': conditions.append("r.year = %s"); params.append(year)
        if municipality != 'ALL': conditions.append("m.municipality_name = %s"); params.append(municipality)
        if type_ != 'ALL': conditions.append("a.type_name = %s"); params.append(type_)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cursor.execute(f"""
            SELECT r.year, m.municipality_name, a.type_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where}
            GROUP BY r.year, m.municipality_name, a.type_name
            ORDER BY total DESC
        """, params)
        rows = cursor.fetchall()

        ws = wb.create_sheet("Dashboard KPI")
        ws.append(['Year', 'Municipality', 'Assistance Type', 'Total Requests'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['year'], row['municipality_name'], row['type_name'], row['total']])
        style_data_rows(ws, 2)

    # ── YoY Trends ──
    if 'yoyTrends' in selected_sections:
        cursor.execute("""
            SELECT r.year, a.type_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            GROUP BY r.year, a.type_name
            ORDER BY r.year, total DESC
        """)
        rows = cursor.fetchall()

        ws = wb.create_sheet("YoY Trends")
        ws.append(['Year', 'Assistance Type', 'Total'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['year'], row['type_name'], row['total']])
        style_data_rows(ws, 2)

    # ── Distribution by Assistance Type ──
    if 'distributionByAssistance' in selected_sections:
        f = filters.get('distributionByAssistance', {})
        top_n = int(f.get('top_n', 5))
        type_ = f.get('type', 'ALL')
        year = f.get('year', 'ALL')

        conditions, params = [], []
        if year != 'ALL': conditions.append("r.year = %s"); params.append(year)
        if type_ != 'ALL': conditions.append("a.type_name = %s"); params.append(type_)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cursor.execute(f"""
            SELECT a.type_name AS name, SUM(r.request_count) AS value
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            {where}
            GROUP BY a.type_name
            ORDER BY value DESC
            LIMIT %s
        """, params + [top_n])
        rows = cursor.fetchall()

        ws = wb.create_sheet("Distribution by Type")
        ws.append(['Assistance Type', 'Total Requests'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['name'], row['value']])
        style_data_rows(ws, 2)

    # ── Distribution by Municipality ──
    if 'distributionByMunicipality' in selected_sections:
        f = filters.get('distributionByMunicipality', {})
        year = f.get('year', 'ALL')

        conditions, params = [], []
        if year != 'ALL': conditions.append("r.year = %s"); params.append(year)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cursor.execute(f"""
            SELECT m.municipality_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where}
            GROUP BY m.municipality_name
            ORDER BY total DESC
        """, params)
        rows = cursor.fetchall()

        ws = wb.create_sheet("Distribution by Municipality")
        ws.append(['Municipality', 'Total Requests'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['municipality_name'], row['total']])
        style_data_rows(ws, 2)

    # ── Comparison Chart ──
    if 'comparisonChart' in selected_sections:
        f = filters.get('comparisonChart', {})
        m1 = f.get('municipality_1', 'ALL')
        m2 = f.get('municipality_2', 'ALL')
        type_ = f.get('type', 'ALL')
        year = f.get('year', 'ALL')

        conditions, params = [], []
        conditions.append("m.municipality_name IN (%s, %s)"); params.extend([m1, m2])
        if year != 'ALL': conditions.append("r.year = %s"); params.append(year)
        if type_ != 'ALL': conditions.append("a.type_name = %s"); params.append(type_)
        where = "WHERE " + " AND ".join(conditions)

        cursor.execute(f"""
            SELECT m.municipality_name, a.type_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where}
            GROUP BY m.municipality_name, a.type_name
            ORDER BY a.type_name
        """, params)
        rows = cursor.fetchall()

        ws = wb.create_sheet("Comparison")
        ws.append(['Municipality', 'Assistance Type', 'Total'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['municipality_name'], row['type_name'], row['total']])
        style_data_rows(ws, 2)

    # ── Municipality Drilldown ──
    if 'municipalityDrilldown' in selected_sections:
        f = filters.get('municipalityDrilldown', {})
        municipality = f.get('municipality', 'ALL')
        year = f.get('year', 'ALL')

        conditions, params = [], []
        if municipality != 'ALL': conditions.append("m.municipality_name = %s"); params.append(municipality)
        if year != 'ALL': conditions.append("r.year = %s"); params.append(year)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cursor.execute(f"""
            SELECT a.type_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where}
            GROUP BY a.type_name
            ORDER BY total DESC
        """, params)
        rows = cursor.fetchall()

        ws = wb.create_sheet("Municipality Drilldown")
        ws.append(['Assistance Type', 'Total'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['type_name'], row['total']])
        style_data_rows(ws, 2)

    # ── Top N Rankings ──
    if 'topNRanking' in selected_sections:
        f = filters.get('topNRanking', {})
        top_n = int(f.get('top_n', 5))
        municipality = f.get('municipality', 'ALL')

        conditions, params = [], []
        if municipality != 'ALL': conditions.append("m.municipality_name = %s"); params.append(municipality)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cursor.execute(f"""
            SELECT m.municipality_name, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where}
            GROUP BY m.municipality_name
            ORDER BY total DESC
            LIMIT %s
        """, params + [top_n])
        rows = cursor.fetchall()

        ws = wb.create_sheet("Top N Rankings")
        ws.append(['Rank', 'Municipality', 'Total Requests'])
        style_header_row(ws, 1)
        for i, row in enumerate(rows):
            ws.append([i + 1, row['municipality_name'], row['total']])
        style_data_rows(ws, 2)

    # ── Forecast ──
    if 'forecast' in selected_sections:
        f = filters.get('forecast', {})
        municipality = f.get('municipality', 'ALL')
        type_ = f.get('type', 'ALL')

        conditions, params = [], []
        if municipality != 'ALL': conditions.append("m.municipality_name = %s"); params.append(municipality)
        if type_ != 'ALL': conditions.append("a.type_name = %s"); params.append(type_)
        where = "WHERE " + " AND ".join(conditions) if conditions else ""

        cursor.execute(f"""
            SELECT r.year, SUM(r.request_count) AS total
            FROM assistance_records r
            JOIN assistance_types a ON r.assistance_type_id = a.type_id
            JOIN municipalities m ON r.municipality_id = m.municipality_id
            {where}
            GROUP BY r.year
            ORDER BY r.year
        """, params)
        rows = cursor.fetchall()

        ws = wb.create_sheet("Forecast Historical")
        ws.append(['Year', 'Total Requests'])
        style_header_row(ws, 1)
        for row in rows:
            ws.append([row['year'], row['total']])
        style_data_rows(ws, 2)

    cursor.close()
    conn.close()

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='MARDSS_Charts.xlsx'
    )